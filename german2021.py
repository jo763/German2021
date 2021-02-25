from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from easygui import choicebox, buttonbox, multenterbox
import random

def exFetch(row, col):
    '''
    Quicker way to extract the contents of a cell from the Excel file
    '''
    cellValue = ws.cell(row = row, column = col).value
    return cellValue


def memoryGame(list1, list2):
    '''
    The main flashcard application
    Shows one side of the card from a list, and then shows the other side

    '''
    lowbound,highbound = rangeSet(list1, list2)

    print("Type 'numchange' and press 'enter' to change the boundaries")
    print("Press 'enter' for the next flashcard")
    while True:
        random_number =random.randint(lowbound,highbound)
        print("\n" + "="*30)
        print(list1[random_number])
        q=input("") #slightly broken code but yolo
        print(list2[random_number])

        p=input("")
        if p == "numchange" or q == "numchange":
          lowbound,highbound =  rangeSet(list1, list2)



def rangeSet(list1, list2):
    '''
    Opens a gui box to get to user to select the

    '''
    # Absolute upper and lower bounds
    abs_lb = 0
    abs_ub = len(list1) - 1
    # Creation of the GUI
    msg = f"Please select the range of words you want to use. If there's a maximim of 10 words in the list, and you select '0' and '4', the first only the first 5 words will appear. \nThe absolute lowerest and highest choices for the list selected are: {abs_lb} and {abs_ub} respectively\nPlease type in your selection. If you don't, the whole list will be used"
    title = "Bound Selection for the box"
    fieldNames = ["Lower Bound", "Upper Bound"]
    fieldValues =[]
    fieldValues = multenterbox(msg,title, fieldNames)

    # Record of user upp and lower bounds
    user_lb = fieldValues[0]
    user_ub = fieldValues[1]

    # Check to see if user choices are valid, if not return the max range,
    try:
        user_ub = int(user_ub)
        user_lb = int(user_lb)
    except:
        return abs_lb, abs_ub

    if user_ub <= abs_ub and user_lb >= abs_lb and abs_ub >= user_lb:
        return user_lb, user_ub
    else:
        return abs_lb, abs_ub


# Name of the input file
filename = "germanContent.xlsx"
# Loads in the file
wb = load_workbook(filename = filename)

# Runs through the sheets in the excel file, creates a popup box asking the user to select their topic
msg ="Please selection a topic"
title = "Topic Selection Box"
choices = wb.sheetnames
sht = choicebox(msg, title, choices)

# Box appears with two buttons asking the user which language they want to see first
msg = "Please select which language you would like to see first"
choices = ["German \n(or list 1)", "English \n(or list 2)"]
title = "Language Order Selection"
LanguageFirstReply = buttonbox(msg, title, choices = choices)
#sht = "Air Travel Nouns"

# Defining the active worksheet
ws = wb[sht]


# Extracting all the topic information for the German and English and placing within lists
germanList = []
englishList = []
i = 1
while True:
    if exFetch(i,1) == None:
        break
    germanList.append(exFetch(i, 1))
    englishList.append(exFetch(i, 2))
    i+=1

if "German" in LanguageFirstReply:
    memoryGame(germanList, englishList)
else:
    memoryGame(englishList, germanList)
